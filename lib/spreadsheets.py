# Functions related to reading and writing data from spreadsheets

import os #, re, shutil, csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


''' Version 3.0
    Version 2.0 adds the ability to deal with spreadsheets with multiple sheets
    Version 3.0 adds the ability to change the column headers and deal with non-unique column headers in spreadsheet
'''

def getSpreadsheetValues(filename, sheetname = ""):
    ''' Gets spreadsheet by name and returns the specified sheet from the spreadsheet as a worksheet. 
        If no sheet is specified it defaults to returning the first sheet '''
    #path = os.path.join('data', filename) 
    wb = load_workbook(filename)
    
    if sheetname != "":
        sheet = wb[sheetname]
    else:
        sheet = wb.worksheets[0]
    values={}

    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    header_map = getColumnHeaders(headers)
    
    for index, col in enumerate(sheet.columns):
        #column = [cell.value for cell in col if cell.value is not None]
        column = [cell.value if cell.value is not None else "" for cell in col]
        
        if len(column) > 0 and column.count("") != len(column):

            values[list(header_map.keys())[index]] = column[1:]
            
    return (values, header_map)

def getColumnHeaders(input_headers):
    ''' Get a list of unique headers for use in the spreadsheet processing, dealing with any repetition in the headings so that all heading are unique
    
        Keyword arguments:
        input_headers -- list of headers from the spreadsheet being processed

        return dictionary: with mapping between original headings (values) and headings to be used during the processing (keys). If all the headings are unique then both key and value will be the same
    '''
    if len(input_headers) == len(set(input_headers)):
        # unique headings so keys and values are the same
        return dict(zip(input_headers, input_headers))
    else:        
        # non-unique headers so renaming headers which are repeated
        new_headings = [input_headers[i] if input_headers.count(input_headers[i]) < 2 else str(input_headers[i]) + "_" + str(input_headers[:i+1].count(input_headers[i])) for i in range(len(input_headers))]
        return dict(zip(new_headings, input_headers))


def replaceColumnHeaders(existing_mapping, new_headers):
    ''' Changes the names of the output column headers in the given mapping to those listed in the given list '''
    count = 0
    newMapping = {}

    for unique_heading in existing_mapping.keys():
        if len(new_headers) > count:
            newMapping[unique_heading] = new_headers[count]
        else:
            newMapping[unique_heading] = str(count + 1)
            print("Warning: replacement headers less than number of columns - column count used for remaining heading")
        count += 1

    if len(new_headers) > count:
        print("Warning: replacement headers greater than number of columns - trailing columns ignored")       

    return newMapping
    

def getSheetListByFilename(filename):
    ''' Get list of sheets in the specified workbook '''
    wb = load_workbook(filename)
    return wb.sheetnames

def getFileList(myDir):
    ''' Get a list of xlsx files in the given directory '''
    return [file for file in myDir.glob("[!~.]*.xlsx")]


def createSheetWithValues(workbook, values, heading_mapping, newValues = {}, sheetname = '', filter = False, min = ''):
    ''' Create a new sheet within the specified workbook with the supplied values (columns in newValues with the same title replace the original version
        in values, can also use filter columns to replace values within a column rather than an entire column)'''

    if sheetname == '':
        newSheet = workbook.active
    else:
        newSheet = workbook[sheetname]
    
    col = 1
    maxRow = 0
    
    #print(values)
    #print(newValues)
    
    for title, column in values.items():
        newSheet.cell(1, col, heading_mapping[title]).font = Font(bold=True)
    
        row = 2
        
        if title in newValues.keys():
            column = newValues[title]
            #print("Length of " + title + " column: " + str(len(column)))

        if filter:
            filteredColumn = zip(column, filter)           

            for filteredRow in filteredColumn:
                #print(str(row[1]) + ": " + str(x) + ", " + str(y))
                # To do: this needs fixing as setting a minimum value won't work
                if (min and filteredRow[1]) or not min:
                    newSheet.cell(row, col, filteredRow[0])
                    row+=1
        else:
            for cell in column:
                newSheet.cell(row, col, cell)
                row+=1  

        if row > maxRow:
            maxRow = row

        #print(title + ": " + str(row))    
        col+=1 

    #print(maxRow)
    #if maxRow > 2:   
    return workbook

def setupNewWorkbook(filename):
    ''' Create and return a workbook that matches the original spreadsheet'''

    original_wb = load_workbook(filename)

    return createWorkbook(original_wb.sheetnames) 



def createWorkbook(sheets = []):
    ''' Create and return a new workbook with the specified sheets'''
    wb = Workbook()

    if len(sheets) < 1:
        return wb
    else:
        for sheetname in sheets:
            wb.create_sheet(sheetname)

        current_sheets = wb.sheetnames
        extras = [item for item in current_sheets if item not in sheets]

        for extra in extras:
            wb.remove(wb[extra])

    return wb


def saveWorkbook(workbook, path, filename, filenameExtra = ''):
    if not os.path.exists(path):
        os.makedirs(path)

    if filenameExtra != '':
        filenameExtra = "_" + str(filenameExtra)

    newFilename = os.path.splitext(os.path.basename(filename))[0] + str(filenameExtra) + os.path.splitext(os.path.basename(filename))[1]
    newFile = os.path.join(path, newFilename)  
    workbook.save(newFile)

    print("New file " + newFile)